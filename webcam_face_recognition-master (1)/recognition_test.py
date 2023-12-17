import face_recognition
import os
import cv2
import numpy as np
import math
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
import sys

def face_confidence(face_distance, face_match_threshold=0.6):
    range_val = (1.0 - face_match_threshold)
    linear_val = (1.0 - face_distance) / (range_val * 2.0)

    if face_distance > face_match_threshold:
        return str(round(linear_val * 100, 2)) + '%'
    else:
        value = (linear_val + ((1.0 - linear_val) * math.pow((linear_val - 0.5) * 2, 0.2))) * 100
        return str(round(value, 2)) + '%'


class FaceRecognition:
    face_locations = []
    face_encodings = []
    face_names = []
    known_face_encodings = []
    known_face_names = []
    process_current_frame = True

    def __init__(self):
        self.encode_faces()
        self.create_attendance_file()

    def encode_faces(self):
        for image in os.listdir('faces'):
            face_image = face_recognition.load_image_file(f"faces/{image}")
            face_encoding = face_recognition.face_encodings(face_image)[0]

            self.known_face_encodings.append(face_encoding)
            self.known_face_names.append(image[:-4])  # Remove file extension
        print(self.known_face_names)

    def create_attendance_file(self):
        filename = 'Attendance.xlsx'

        if not os.path.exists(filename):
            workbook = Workbook()
            worksheet = workbook.active

            # Set the headers
            worksheet.cell(row=1, column=1, value='Names')

            for i, name in enumerate(self.known_face_names, start=2):
                worksheet.cell(row=i, column=1, value=name)

            # Set the dates
            current_date = datetime.datetime.now().date()  # Use datetime.date() instead of datetime.datetime()
            num_dates = 31  # Set the number of dates to be displayed

            for i in range(num_dates):
                date = current_date + datetime.timedelta(days=i)
                worksheet.cell(row=1, column=i + 2, value=date.strftime('%d-%m-%Y'))  # Use the correct date format

            # Set the total column
            worksheet.cell(row=1, column=num_dates + 2, value='Total')

            workbook.save(filename)
            print(f"Attendance file created: {filename}")

    def get_date_column(self, worksheet, date):
        formatted_date = date.strftime('%d-%m-%Y')
        for column in range(2, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=column)
            cell_date_str = cell.value.strftime('%d-%m-%Y')  # Convert cell value to the correct date format
            if cell_date_str == formatted_date:
                return column
        return None


    def mark_attendance(self, name):
        filename = 'Attendance.xlsx'
        workbook = load_workbook(filename)
        worksheet = workbook['Sheet']

        current_date = datetime.date.today()  # Ensure current_date is a datetime.date object
        current_time = datetime.datetime.now().strftime("%H:%M:%S")

        if name in self.known_face_names:
            student_row = self.known_face_names.index(name) + 2
            date_column = self.get_date_column(worksheet, current_date)

            if date_column is not None:
                cell = worksheet.cell(row=student_row, column=date_column)
                if cell.value == 'P':
                    print(f"{name} is already marked present for {current_date.strftime('%d-%m-%Y')}")
                else:
                    cell.value = 'P'
                    comment = Comment(f"Time: {current_time}", "Attendance System")
                    cell.comment = comment
                    print(f"Marked {name} as present for {current_date.strftime('%d-%m-%Y')} at {current_time}")
            else:
                print(f"No column found for today's date: {current_date.strftime('%d-%m-%Y')}")
        else:
            print(f"Unknown student: {name}")

        workbook.save(filename)

    def run_recognition(self):
        video_capture = cv2.VideoCapture(0)
        #video_capture = cv2.VideoCapture('http://192.168.131.29:81/stream')

        if not video_capture.isOpened():
            sys.exit('Video source not found...')

        while True:
            ret, frame = video_capture.read()

            # Only process every other frame of video to save time
            if self.process_current_frame:
                # Resize frame of video to 1/4 size for faster face recognition processing
                small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

                # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
                rgb_small_frame = small_frame[:, :, ::-1]

                # Find all the faces and face encodings in the current frame of video
                self.face_locations = face_recognition.face_locations(rgb_small_frame)
                self.face_encodings = face_recognition.face_encodings(rgb_small_frame, self.face_locations)

                self.face_names = []
                for face_encoding in self.face_encodings:
                    # See if the face is a match for the known face(s)
                    matches = face_recognition.compare_faces(self.known_face_encodings, face_encoding)
                    name = "Unknown"
                    confidence = '???'

                    # Calculate the shortest distance to face
                    face_distances = face_recognition.face_distance(self.known_face_encodings, face_encoding)

                    best_match_index = np.argmin(face_distances)
                    if matches[best_match_index]:
                        name = self.known_face_names[best_match_index]
                        confidence = face_confidence(face_distances[best_match_index])

                    print(name)
                    #self.mark_attendance(name)
                    #self.face_names.append(f'{name} ({confidence})')

            self.process_current_frame = not self.process_current_frame

            # Display the results
            for (top, right, bottom, left), name in zip(self.face_locations, self.face_names):
                # Scale back up face locations since the frame we detected in was scaled to 1/4 size
                top *= 4
                right *= 4
                bottom *= 4
                left *= 4

                # Create the frame with the name
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
                cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 1)

            # Display the resulting image
            cv2.imshow('Face Recognition', frame)

            # Hit 'q' on the keyboard to quit!
            if cv2.waitKey(1) == ord('q'):
                break

        # Release handle to the webcam
        video_capture.release()
        cv2.destroyAllWindows()


if __name__ == '__main__':
    fr = FaceRecognition()
    fr.run_recognition()
